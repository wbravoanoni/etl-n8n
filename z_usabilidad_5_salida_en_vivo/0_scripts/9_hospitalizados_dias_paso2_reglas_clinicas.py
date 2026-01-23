import logging
import math
from datetime import datetime, date, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# CONFIGURACIÓN LOGS
# =========================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# =========================================================
# ARCHIVOS
# =========================================================
archivo_entrada = (
    "z_usabilidad_5_salida_en_vivo/1_entrada/"
    "9_hospitalizados_dias_paso1_descarga.xlsx"
)

archivo_salida = (
    "z_usabilidad_5_salida_en_vivo/3_resultados/"
    "9_hospitalizados_dias_paso2_reglas_clinicas.xlsx"
)

MINUTOS_DIA = 1440
now = datetime.now()

# =========================================================
# FUNCIONES AUXILIARES
# =========================================================
def parse_fecha(v):
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v), "%Y-%m-%d").date()
    except:
        return None

def parse_hora(v):
    if isinstance(v, time):
        return v
    try:
        return datetime.strptime(str(v), "%H:%M:%S").time()
    except:
        return None

def build_dt(f, h):
    return datetime.combine(f, h) if f and h else None

def minutos(inicio, fin):
    if not inicio or not fin:
        return 0
    return int((fin - inicio).total_seconds() / 60)

def dias_admin(mins):
    # Regla administrativa: cualquier minuto cuenta como 1 día
    return math.ceil(mins / MINUTOS_DIA) if mins > 0 else 0

# =========================================================
# PROCESO PRINCIPAL
# =========================================================
try:
    wb = load_workbook(archivo_entrada)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    wb_out = Workbook()
    ws_out = wb_out.active

    ws_out.append(headers + [
        "minutos_estadia_servicio",
        "dias_estadia_servicio",
        "minutos_estadia_episodio",
        "dias_estadia_episodio"
    ])

    for c in ws_out[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(min_row=2, values_only=True):

        # -------------------------------------------------
        # SERVICIO
        # -------------------------------------------------
        ini_srv = build_dt(
            parse_fecha(row[idx["fecha_inicio_servicio"]]),
            parse_hora(row[idx["hora_inicio_servicio"]])
        )

        fin_srv_raw = build_dt(
            parse_fecha(row[idx["fecha_termino_servicio"]]),
            parse_hora(row[idx["hora_termino_servicio"]])
        )

        # Si no hay término, se usa now SOLO para cálculo
        fin_srv_calc = fin_srv_raw or now

        min_srv = minutos(ini_srv, fin_srv_calc)
        dias_srv = dias_admin(min_srv)

        # -------------------------------------------------
        # EPISODIO
        # -------------------------------------------------
        ini_epi = build_dt(
            parse_fecha(row[idx["fecha_admision"]]),
            parse_hora(row[idx["hora_admision"]])
        )

        fin_epi_raw = build_dt(
            parse_fecha(row[idx["fechaAltaAdm"]]),
            parse_hora(row[idx["horaAltaAdm"]])
        )

        fin_epi_calc = fin_epi_raw or now

        min_epi = minutos(ini_epi, fin_epi_calc)
        dias_epi = dias_admin(min_epi)

        ws_out.append(list(row) + [
            min_srv,
            dias_srv,
            min_epi,
            dias_epi
        ])

    for col in ws_out.columns:
        ws_out.column_dimensions[col[0].column_letter].width = 30

    wb_out.save(archivo_salida)
    logging.info(f"Archivo generado correctamente: {archivo_salida}")

except Exception as e:
    logging.error(f"Error en PASO 2: {e}")
