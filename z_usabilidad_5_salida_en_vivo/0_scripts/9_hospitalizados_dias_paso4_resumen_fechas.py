import logging
import math
from datetime import datetime, date, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# =====================================================
# CONFIGURACIÓN LOGS
# =====================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# =====================================================
# ARCHIVOS
# =====================================================
archivo_entrada = (
    "z_usabilidad_5_salida_en_vivo/3_resultados/"
    "9_hospitalizados_dias_paso2_reglas_clinicas.xlsx"
)

archivo_salida = (
    "z_usabilidad_5_salida_en_vivo/3_resultados/"
    "9_hospitalizados_dias_paso4_resumen_servicios.xlsx"
)

MINUTOS_DIA = 1440
now = datetime.now()

# =====================================================
# FUNCIONES AUXILIARES
# =====================================================
def parse_fecha(v):
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v), "%Y-%m-%d").date()
    except:
        return None

def parse_hora(v):
    if isinstance(v, time):
        return v
    try:
        return datetime.strptime(str(v), "%H:%M:%S").time()
    except:
        return None

def build_dt(f, h):
    return datetime.combine(f, h) if f and h else None

def minutos(inicio, fin):
    if not inicio or not fin:
        return 0
    return int((fin - inicio).total_seconds() / 60)

def dias_admin(mins):
    # Regla administrativa: cualquier minuto cuenta como día
    return math.ceil(mins / MINUTOS_DIA) if mins > 0 else 0

# =====================================================
# PROCESO PRINCIPAL
# =====================================================
try:
    wb = load_workbook(archivo_entrada)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    logging.info(f"Filas leídas: {len(filas)}")

    episodios = {}

    # -------------------------------------------------
    # AGRUPACIÓN POR EPISODIO
    # -------------------------------------------------
    for r in filas:
        episodio = r[idx["episodio"]]

        if episodio not in episodios:
            episodios[episodio] = {
                "EstadoAtencion": r[idx["EstadoAtencion"]],
                "fecha_admision": parse_fecha(r[idx["fecha_admision"]]),
                "hora_admision": parse_hora(r[idx["hora_admision"]]),
                "fechaAltaAdm": parse_fecha(r[idx["fechaAltaAdm"]]),
                "horaAltaAdm": parse_hora(r[idx["horaAltaAdm"]]),
                "ini_srv": None,
                "fin_srv": None
            }

        # -------- SERVICIO --------
        ini_srv = build_dt(
            parse_fecha(r[idx["fecha_inicio_servicio"]]),
            parse_hora(r[idx["hora_inicio_servicio"]])
        )

        fin_srv_raw = build_dt(
            parse_fecha(r[idx["fecha_termino_servicio"]]),
            parse_hora(r[idx["hora_termino_servicio"]])
        )

        fin_srv = fin_srv_raw or now

        if ini_srv:
            if episodios[episodio]["ini_srv"] is None:
                episodios[episodio]["ini_srv"] = ini_srv
            else:
                episodios[episodio]["ini_srv"] = min(
                    episodios[episodio]["ini_srv"],
                    ini_srv
                )

        if episodios[episodio]["fin_srv"] is None:
            episodios[episodio]["fin_srv"] = fin_srv
        else:
            episodios[episodio]["fin_srv"] = max(
                episodios[episodio]["fin_srv"],
                fin_srv
            )

    logging.info(f"Episodios únicos: {len(episodios)}")

    # =====================================================
    # ESCRITURA ARCHIVO FINAL
    # =====================================================
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "resumen_episodio"

    ws_out.append([
        "episodio",
        "EstadoAtencion",
        "fecha_admision_completa",
        "fechaAltaAdm_completa",
        "fecha_inicio_servicio_completo",
        "fecha_termino_servicio_completo",
        "minutos_estadia_episodio",
        "dias_estadia_episodio",
        "minutos_estadia_servicio",
        "dias_estadia_servicio",
        "comparacion_fechas"
    ])

    for cell in ws_out[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # -------------------------------------------------
    # CÁLCULOS FINALES
    # -------------------------------------------------
    for episodio, d in episodios.items():

        ini_epi = build_dt(d["fecha_admision"], d["hora_admision"])
        fin_epi = build_dt(d["fechaAltaAdm"], d["horaAltaAdm"]) or now

        ini_srv = d["ini_srv"]
        fin_srv = d["fin_srv"]

        min_epi = minutos(ini_epi, fin_epi)
        dias_epi = dias_admin(min_epi)

        min_srv = minutos(ini_srv, fin_srv)
        dias_srv = dias_admin(min_srv)

        comparacion = "si" if dias_epi == dias_srv else "no"

        ws_out.append([
            episodio,
            d["EstadoAtencion"],
            ini_epi,
            fin_epi if d["fechaAltaAdm"] else None,
            ini_srv,
            fin_srv if fin_srv != now else None,
            min_epi,
            dias_epi,
            min_srv,
            dias_srv,
            comparacion
        ])

    for col in ws_out.columns:
        ws_out.column_dimensions[col[0].column_letter].width = 30

    wb_out.save(archivo_salida)
    logging.info(f"Archivo generado correctamente: {archivo_salida}")

except Exception as e:
    logging.error(f"Error en resumen de servicios: {e}")
