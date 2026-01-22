import logging
import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# =========================================================
# CONFIGURACIÓN LOGS
# =========================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# =========================================================
# ARCHIVOS
# =========================================================
archivo_entrada = (
    "z_usabilidad_5_salida_en_vivo/3_resultados/"
    "9_hospitalizados_dias_pro_consolidado.xlsx"
)

archivo_salida = (
    "z_usabilidad_5_salida_en_vivo/3_resultados/"
    "9_hospitalizados_dias_pro2.xlsx"
)

MINUTOS_DIA = 1440

try:
    # =====================================================
    # LECTURA ARCHIVO BASE
    # =====================================================
    wb = load_workbook(archivo_entrada)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    logging.info(f"Filas originales: {len(filas)}")

    # =====================================================
    # AGRUPACIÓN POR EPISODIO
    # =====================================================
    resumen = {}

    for row in filas:
        episodio = row[idx["episodio"]]

        min_srv = row[idx["minutos_estadia_servicio_sum"]]
        min_epi = row[idx["minutos_estadia_episodio"]]

        if episodio not in resumen:
            resumen[episodio] = {
                "minutos_servicio": 0,
                "minutos_episodio": min_epi
            }

        if isinstance(min_srv, (int, float)):
            resumen[episodio]["minutos_servicio"] += min_srv

    logging.info(f"Episodios únicos: {len(resumen)}")

    # =====================================================
    # ESCRITURA RESULTADO FINAL
    # =====================================================
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "resumen_episodio"

    ws_out.append([
        "episodio",
        "minutos_estadia_servicio_sum",
        "dias_estadia_servicio_sum",
        "minutos_estadia_episodio",
        "dias_estadia_episodio"
    ])

    for cell in ws_out[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for episodio, data in resumen.items():

        min_srv_total = data["minutos_servicio"]
        min_epi_total = data["minutos_episodio"]

        dias_srv = (
            math.ceil(min_srv_total / MINUTOS_DIA)
            if min_srv_total > 0 else 0
        )

        dias_epi = (
            math.ceil(min_epi_total / MINUTOS_DIA)
            if min_epi_total > 0 else 0
        )

        ws_out.append([
            episodio,
            min_srv_total,
            dias_srv,
            min_epi_total,
            dias_epi
        ])

    for col in ws_out.columns:
        ws_out.column_dimensions[col[0].column_letter].width = 32

    wb_out.save(archivo_salida)
    logging.info(f"Archivo generado correctamente: {archivo_salida}")

except Exception as e:
    logging.error(f"Error en la prueba de agrupación: {e}")
