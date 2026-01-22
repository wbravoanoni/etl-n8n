import logging
import math
from datetime import datetime, date, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

archivo_entrada = "z_usabilidad_5_salida_en_vivo/3_resultados/9_hospitalizados_dias_pro.xlsx"
archivo_salida  = "z_usabilidad_5_salida_en_vivo/3_resultados/9_hospitalizados_dias_pro_consolidado.xlsx"

MINUTOS_DIA = 1440
now = datetime.now()

# =======================
# Funciones auxiliares
# =======================
def parse_fecha(v):
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v), "%Y-%m-%d").date()
    except:
        return None


def parse_hora(v):
    if isinstance(v, time):
        return v
    try:
        return datetime.strptime(str(v), "%H:%M:%S").time()
    except:
        return None


def build_dt(f, h):
    return datetime.combine(f, h) if f and h else None


def minutos(i, f):
    if not i or not f:
        return 0
    return int((f - i).total_seconds() / 60)


def dias_cama(mins):
    # Regla clínica-administrativa:
    # cualquier minuto cuenta como día
    return math.ceil(mins / MINUTOS_DIA) if mins > 0 else 0


# =======================
# Proceso principal
# =======================
try:
    wb = load_workbook(archivo_entrada)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "hospitalizados_consolidado"

    ws_out.append(headers + [
        "minutos_estadia_servicio_sum",
        "dias_estadia_servicio_sum",
        "minutos_estadia_episodio",
        "dias_estadia_episodio",
        "flag_servicio_abierto"
    ])

    for cell in ws_out[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    grupos = {}

    # -----------------------
    # Consolidación
    # -----------------------
    for r in ws.iter_rows(min_row=2, values_only=True):

        episodio = r[idx["episodio"]]
        servicio = r[idx["servicio"]]
        fecha_ini_srv = parse_fecha(r[idx["fecha_inicio_servicio"]])

        ini_srv = build_dt(
            fecha_ini_srv,
            parse_hora(r[idx["hora_inicio_servicio"]])
        )

        fin_srv_real = build_dt(
            parse_fecha(r[idx["fecha_termino_servicio"]]),
            parse_hora(r[idx["hora_termino_servicio"]])
        )

        servicio_abierto = fin_srv_real is None
        fin_srv_calc = fin_srv_real if fin_srv_real else now

        key = (episodio, servicio, fecha_ini_srv, servicio_abierto)

        if key not in grupos:
            grupos[key] = {
                "row": list(r),
                "ini": ini_srv,
                "fin_calc": fin_srv_calc,
                "fin_real": fin_srv_real,
                "abierto": servicio_abierto
            }
        else:
            grupos[key]["ini"] = min(grupos[key]["ini"], ini_srv)
            grupos[key]["fin_calc"] = max(grupos[key]["fin_calc"], fin_srv_calc)

    # -----------------------
    # Escritura final
    # -----------------------
    for g in grupos.values():
        r = g["row"]

        ini_srv = g["ini"]
        fin_calc = g["fin_calc"]
        fin_real = g["fin_real"]
        abierto = g["abierto"]

        # inicio siempre visible
        r[idx["fecha_inicio_servicio"]] = ini_srv.date()
        r[idx["hora_inicio_servicio"]]  = ini_srv.time()

        # 👇 CLAVE: NO escribir fecha término si está abierto
        if not abierto:
            r[idx["fecha_termino_servicio"]] = fin_real.date()
            r[idx["hora_termino_servicio"]]  = fin_real.time()
        else:
            r[idx["fecha_termino_servicio"]] = None
            r[idx["hora_termino_servicio"]]  = None

        # ---- servicio ----
        min_srv = minutos(ini_srv, fin_calc)
        dias_srv = dias_cama(min_srv)

        # ---- episodio ----
        ini_epi = build_dt(
            parse_fecha(r[idx["fecha_admision"]]),
            parse_hora(r[idx["hora_admision"]])
        )

        fin_epi = build_dt(
            parse_fecha(r[idx["fechaAltaAdm"]]),
            parse_hora(r[idx["horaAltaAdm"]])
        ) or now

        min_epi = minutos(ini_epi, fin_epi)
        dias_epi = dias_cama(min_epi)

        ws_out.append(r + [
            min_srv,
            dias_srv,
            min_epi,
            dias_epi,
            "SI" if abierto else "NO"
        ])

    for col in ws_out.columns:
        ws_out.column_dimensions[col[0].column_letter].width = 25

    wb_out.save(archivo_salida)
    logging.info(f"Archivo consolidado generado correctamente: {archivo_salida}")

except Exception as e:
    logging.error(f"Error en consolidación: {e}")
