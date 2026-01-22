import logging
from datetime import datetime, date, time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

archivo_entrada = "z_usabilidad_5_salida_en_vivo/1_entrada/9_hospitalizados_dias.xlsx"
archivo_salida  = "z_usabilidad_5_salida_en_vivo/3_resultados/9_hospitalizados_dias_pro.xlsx"

# =========================================================
# FUNCIONES AUXILIARES
# =========================================================
def parse_fecha(v):
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v), "%Y-%m-%d").date()
    except Exception:
        return None


def parse_hora(v):
    if isinstance(v, time):
        return v
    try:
        return datetime.strptime(str(v), "%H:%M:%S").time()
    except Exception:
        return None


def build_dt(f, h):
    return datetime.combine(f, h) if f and h else None


# =========================================================
# PROCESO PRINCIPAL
# =========================================================
try:
    wb = load_workbook(archivo_entrada)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.append(headers)

    now = datetime.now()

    for row in ws.iter_rows(min_row=2, values_only=True):

        # -------------------------------------------------
        # SERVICIO
        # -------------------------------------------------
        inicio_srv = build_dt(
            parse_fecha(row[idx["fecha_inicio_servicio"]]),
            parse_hora(row[idx["hora_inicio_servicio"]])
        )

        fin_srv_raw = build_dt(
            parse_fecha(row[idx["fecha_termino_servicio"]]),
            parse_hora(row[idx["hora_termino_servicio"]])
        )

        # REGLA NUEVA (CLÍNICA):
        # Servicio sin fecha de término → se considera activo
        # hasta la fecha/hora de ejecución del script
        if inicio_srv and not fin_srv_raw:
            fin_srv = now
        else:
            fin_srv = fin_srv_raw

        # -------------------------------------------------
        # EPISODIO
        # -------------------------------------------------
        inicio_epi = build_dt(
            parse_fecha(row[idx["fecha_admision"]]),
            parse_hora(row[idx["hora_admision"]])
        )

        fin_epi = build_dt(
            parse_fecha(row[idx["fechaAltaAdm"]]),
            parse_hora(row[idx["horaAltaAdm"]])
        ) or now

        # -------------------------------------------------
        # REGLA 1 (YA EXISTENTE)
        # -------------------------------------------------
        if inicio_srv and fin_srv and inicio_epi and inicio_srv > fin_srv and inicio_epi == fin_srv:
            continue

        # -------------------------------------------------
        # REGLA 2 (YA EXISTENTE)
        # -------------------------------------------------
        if inicio_epi and fin_epi and inicio_srv and inicio_epi > fin_epi:
            inicio_epi = inicio_srv

        # -------------------------------------------------
        # SE MANTIENE EL REGISTRO
        # -------------------------------------------------
        ws_out.append(row)

    wb_out.save(archivo_salida)
    logging.info(f"Archivo generado correctamente: {archivo_salida}")

except Exception as e:
    logging.error(f"Error en preprocesamiento: {e}")
