import logging
import math
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment

# =====================================================
# CONFIGURACIÓN LOGS
# =====================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# =====================================================
# ARCHIVOS
# =====================================================
archivo_entrada = (
    "z_usabilidad_5_salida_en_vivo/3_resultados/"
    "9_hospitalizados_dias_paso4_resumen_servicios.xlsx"
)

archivo_salida = (
    "z_usabilidad_5_salida_en_vivo/3_resultados/"
    "9_hospitalizados_dias_paso5_resumen_ajustado.xlsx"
)

MINUTOS_DIA = 1440
now = datetime.now()

# =====================================================
# FUNCIONES
# =====================================================
def minutos(inicio, fin):
    if not inicio or not fin:
        return 0
    return int((fin - inicio).total_seconds() / 60)

def dias_admin(mins):
    return math.ceil(mins / MINUTOS_DIA) if mins > 0 else 0

# =====================================================
# PROCESO PRINCIPAL
# =====================================================
try:
    wb = load_workbook(archivo_entrada)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "resumen_ajustado"

    ws_out.append(headers)

    for cell in ws_out[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    logging.info(f"Filas leídas: {len(filas)}")

    # =================================================
    # AJUSTE + RECÁLCULO
    # =================================================
    for r in filas:
        r = list(r)

        ini_epi = r[idx["fecha_admision_completa"]]
        fin_epi = r[idx["fechaAltaAdm_completa"]] or now

        ini_srv = r[idx["fecha_inicio_servicio_completo"]]
        fin_srv = r[idx["fecha_termino_servicio_completo"]] or now

        # ---------------------------------------------
        # AJUSTE SOLO SI comparacion_fechas = "no"
        # ---------------------------------------------
        if r[idx["comparacion_fechas"]] == "no":
            if ini_srv and ini_epi and ini_srv < ini_epi:
                ini_srv = ini_epi
                r[idx["fecha_inicio_servicio_completo"]] = ini_srv

        # ---------------------------------------------
        # RECÁLCULO EPISODIO
        # ---------------------------------------------
        min_epi = minutos(ini_epi, fin_epi)
        dias_epi = dias_admin(min_epi)

        # ---------------------------------------------
        # RECÁLCULO SERVICIO
        # ---------------------------------------------
        min_srv = minutos(ini_srv, fin_srv)
        dias_srv = dias_admin(min_srv)

        # ---------------------------------------------
        # ACTUALIZAR COLUMNAS
        # ---------------------------------------------
        r[idx["minutos_estadia_episodio"]] = min_epi
        r[idx["dias_estadia_episodio"]] = dias_epi
        r[idx["minutos_estadia_servicio"]] = min_srv
        r[idx["dias_estadia_servicio"]] = dias_srv

        r[idx["comparacion_fechas"]] = (
            "si" if dias_epi == dias_srv else "no"
        )

        ws_out.append(r)

    for col in ws_out.columns:
        ws_out.column_dimensions[col[0].column_letter].width = 30

    wb_out.save(archivo_salida)
    logging.info(f"Archivo ajustado generado correctamente: {archivo_salida}")

except Exception as e:
    logging.error(f"Error en ajuste de comparación: {e}")
